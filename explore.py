"""Exploratory analysis: read merged data, compute correlations, generate charts."""
# pylint: disable=line-too-long,wrong-import-position,invalid-name,unsubscriptable-object,use-dict-literal
import os
from pathlib import Path

ROOT = Path(__file__).resolve().parent
CACHE_DIR = ROOT / ".cache"
CACHE_DIR.mkdir(exist_ok=True)
(CACHE_DIR / "matplotlib").mkdir(exist_ok=True)
os.environ.setdefault("MPLBACKEND", "Agg")
os.environ.setdefault("MPLCONFIGDIR", str(CACHE_DIR / "matplotlib"))
os.environ.setdefault("XDG_CACHE_HOME", str(CACHE_DIR))

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

XLSX = ROOT / "data" / "AI_Adoption_Research_Data.xlsx"
OUTDIR = ROOT / "figures"
OUTDIR.mkdir(exist_ok=True)

# Use a clean style
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 10,
    "axes.titleweight": "bold",
    "axes.titlesize": 13,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.grid": True,
    "grid.alpha": 0.3,
    "figure.dpi": 130,
})

# Read calculated values from the workbook (data_only=True so formulas resolve)
wb = load_workbook(XLSX, data_only=True)

# ----- Master country -----
ws = wb["Master_Country"]
rows = list(ws.iter_rows(min_row=5, values_only=True))
master = pd.DataFrame(rows, columns=["Country", "Adoption", "Optimism", "GDP_PPP", "Internet", "Tertiary"])
print("Master country sheet:")
print(master.to_string(index=False))
print(f"\nN with adoption data: {master['Adoption'].notna().sum()}")
print(f"N with optimism data: {master['Optimism'].notna().sum()}")

# ----- Correlations using only countries with full data -----
adopt = master.dropna(subset=["Adoption"])
print(f"\n--- Correlations (n={len(adopt)} countries with adoption data) ---")
for col in ["GDP_PPP", "Internet", "Tertiary"]:
    sub = adopt.dropna(subset=[col])
    r = np.corrcoef(sub["Adoption"].astype(float), sub[col].astype(float))[0,1]
    print(f"AI Adoption × {col:<10}: r = {r:+.3f}  (n={len(sub)})")

opt_sub = master.dropna(subset=["Optimism"])
print(f"\n--- Optimism correlations (n={len(opt_sub)}) ---")
for col in ["GDP_PPP", "Internet", "Tertiary"]:
    sub = opt_sub.dropna(subset=[col])
    r = np.corrcoef(sub["Optimism"].astype(float), sub[col].astype(float))[0,1]
    print(f"AI Optimism × {col:<10}: r = {r:+.3f}  (n={len(sub)})")

both = master.dropna(subset=["Adoption", "Optimism"])
r = np.corrcoef(both["Adoption"].astype(float), both["Optimism"].astype(float))[0,1]
print(f"AI Adoption × AI Optimism: r = {r:+.3f}  (n={len(both)})")

# =========================================================
# CHART 1: US Adoption by Age (Pew 2025)
# =========================================================
fig, ax = plt.subplots(figsize=(8, 4.5))
ages = ["18-29", "30-49", "50-64", "65+"]
pcts = [58, 41, 25, 10]
colors = ["#1F4E78", "#2E75B6", "#5B9BD5", "#9DC3E6"]
bars = ax.bar(ages, pcts, color=colors, edgecolor="white", linewidth=1.5)
ax.set_ylabel("% of US adults who have used ChatGPT")
ax.set_title("ChatGPT Adoption by Age — US Adults (Pew Research, 2025)")
ax.set_ylim(0, 70)
for b, p in zip(bars, pcts):
    ax.text(b.get_x() + b.get_width()/2, b.get_height() + 1.5, f"{p}%", ha="center", fontweight="bold")
ax.axhline(34, color="#C00000", linestyle="--", linewidth=1, alpha=0.7)
ax.text(3.4, 35.5, "All adults: 34%", color="#C00000", fontsize=9, ha="right")
fig.tight_layout()
fig.savefig(OUTDIR / "chart1_us_adoption_by_age.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 2: US Adoption by Education
# =========================================================
fig, ax = plt.subplots(figsize=(8, 4.5))
ed = ["HS or less", "Some college", "Bachelor's", "Postgrad"]
pcts = [18, 33, 51, 52]
colors = ["#9DC3E6", "#5B9BD5", "#2E75B6", "#1F4E78"]
bars = ax.bar(ed, pcts, color=colors, edgecolor="white", linewidth=1.5)
ax.set_ylabel("% of US adults who have used ChatGPT")
ax.set_title("ChatGPT Adoption by Education — US Adults (Pew Research, 2025)")
ax.set_ylim(0, 65)
for b, p in zip(bars, pcts):
    ax.text(b.get_x() + b.get_width()/2, b.get_height() + 1.5, f"{p}%", ha="center", fontweight="bold")
fig.tight_layout()
fig.savefig(OUTDIR / "chart2_us_adoption_by_education.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 3: Top 30 countries by AI adoption (Microsoft)
# =========================================================
ms_top = master.dropna(subset=["Adoption"]).sort_values("Adoption", ascending=True)
fig, ax = plt.subplots(figsize=(9, 11))
colors = ["#1F4E78" if v >= 0.40 else "#2E75B6" if v >= 0.30 else "#5B9BD5" for v in ms_top["Adoption"]]
ax.barh(ms_top["Country"], ms_top["Adoption"]*100, color=colors, edgecolor="white", linewidth=0.5)
ax.set_xlabel("% of working-age population using generative AI tools")
ax.set_title("Generative AI Adoption by Country — Microsoft AI Diffusion Report (H2 2025)")
ax.axvline(16.3, color="#C00000", linestyle="--", linewidth=1, alpha=0.7)
ax.text(16.5, -0.5, "Global avg 16.3%", color="#C00000", fontsize=9)
for i, (c, v) in enumerate(zip(ms_top["Country"], ms_top["Adoption"])):
    ax.text(v*100 + 0.5, i, f"{v*100:.1f}%", va="center", fontsize=8)
ax.set_xlim(0, 75)
fig.tight_layout()
fig.savefig(OUTDIR / "chart3_country_adoption.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 4: Scatter — AI Adoption vs GDP per capita
# =========================================================
sub = master.dropna(subset=["Adoption", "GDP_PPP"]).copy()
sub["Adoption"] = sub["Adoption"].astype(float)
sub["GDP_PPP"] = sub["GDP_PPP"].astype(float)
fig, ax = plt.subplots(figsize=(9, 6))
ax.scatter(sub["GDP_PPP"]/1000, sub["Adoption"]*100, s=80, color="#2E75B6", edgecolor="#1F4E78", linewidth=0.8, alpha=0.85)
# Label notable countries
labels_show = ["United Arab Emirates", "Singapore", "Norway", "Ireland", "France", "United States",
               "China", "India", "Japan", "Germany", "Brazil", "Italy"]
for _, r in sub.iterrows():
    if r["Country"] in labels_show:
        ax.annotate(r["Country"], (r["GDP_PPP"]/1000, r["Adoption"]*100),
                    xytext=(5,5), textcoords="offset points", fontsize=8, color="#333")
# Trend line
z = np.polyfit(sub["GDP_PPP"]/1000, sub["Adoption"]*100, 1)
xx = np.linspace(sub["GDP_PPP"].min()/1000, sub["GDP_PPP"].max()/1000, 100)
ax.plot(xx, z[0]*xx + z[1], color="#C00000", linestyle="--", linewidth=1.5, alpha=0.7, label=f"Linear fit: y = {z[0]:.3f}x + {z[1]:.1f}")
r = np.corrcoef(sub["GDP_PPP"], sub["Adoption"])[0,1]
ax.text(0.02, 0.95, f"Pearson r = {r:.3f}\nn = {len(sub)} countries",
        transform=ax.transAxes, fontsize=10, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="#1F4E78"))
ax.set_xlabel("GDP per capita PPP 2024 ($000s)")
ax.set_ylabel("% population using generative AI (H2 2025)")
ax.set_title("AI Adoption vs. GDP per capita — Country-Level")
ax.legend(loc="lower right")
fig.tight_layout()
fig.savefig(OUTDIR / "chart4_adoption_vs_gdp.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 5: Scatter — AI Adoption vs Internet Penetration
# =========================================================
sub = master.dropna(subset=["Adoption", "Internet"]).copy()
sub["Adoption"] = sub["Adoption"].astype(float)
sub["Internet"] = sub["Internet"].astype(float)
fig, ax = plt.subplots(figsize=(9, 6))
ax.scatter(sub["Internet"], sub["Adoption"]*100, s=80, color="#2E75B6", edgecolor="#1F4E78", linewidth=0.8, alpha=0.85)
for _, row in sub.iterrows():
    if row["Country"] in labels_show:
        ax.annotate(row["Country"], (row["Internet"], row["Adoption"]*100),
                    xytext=(5,5), textcoords="offset points", fontsize=8, color="#333")
z = np.polyfit(sub["Internet"], sub["Adoption"]*100, 1)
xx = np.linspace(sub["Internet"].min(), sub["Internet"].max(), 100)
ax.plot(xx, z[0]*xx + z[1], color="#C00000", linestyle="--", linewidth=1.5, alpha=0.7, label=f"Linear fit: y = {z[0]:.3f}x + {z[1]:.1f}")
r = np.corrcoef(sub["Internet"], sub["Adoption"])[0,1]
ax.text(0.02, 0.95, f"Pearson r = {r:.3f}\nn = {len(sub)} countries",
        transform=ax.transAxes, fontsize=10, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="#1F4E78"))
ax.set_xlabel("Internet users (% of population, 2024)")
ax.set_ylabel("% population using generative AI (H2 2025)")
ax.set_title("AI Adoption vs. Internet Penetration — Country-Level")
ax.legend(loc="lower right")
fig.tight_layout()
fig.savefig(OUTDIR / "chart5_adoption_vs_internet.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 6: AI Optimism vs Adoption (the counter-intuitive one)
# =========================================================
sub = master.dropna(subset=["Adoption", "Optimism"]).copy()
sub["Adoption"] = sub["Adoption"].astype(float)
sub["Optimism"] = sub["Optimism"].astype(float)
fig, ax = plt.subplots(figsize=(9, 6))
ax.scatter(sub["Optimism"]*100, sub["Adoption"]*100, s=80, color="#7030A0", edgecolor="#3F1670", linewidth=0.8, alpha=0.85)
for _, row in sub.iterrows():
    ax.annotate(row["Country"], (row["Optimism"]*100, row["Adoption"]*100),
                xytext=(5,5), textcoords="offset points", fontsize=8, color="#333")
z = np.polyfit(sub["Optimism"], sub["Adoption"], 1)
xx = np.linspace(sub["Optimism"].min(), sub["Optimism"].max(), 100)
ax.plot(xx*100, (z[0]*xx + z[1])*100, color="#C00000", linestyle="--", linewidth=1.5, alpha=0.7)
r = np.corrcoef(sub["Optimism"], sub["Adoption"])[0,1]
ax.text(0.02, 0.95, f"Pearson r = {r:+.3f}\nn = {len(sub)} countries",
        transform=ax.transAxes, fontsize=10, fontweight="bold",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="white", edgecolor="#7030A0"))
ax.set_xlabel("% saying AI is more beneficial than harmful (Stanford/Ipsos)")
ax.set_ylabel("% using generative AI (Microsoft Diffusion H2 2025)")
ax.set_title("Counter-Intuitive: Skeptics Adopt More — Country-Level")
fig.tight_layout()
fig.savefig(OUTDIR / "chart6_optimism_vs_adoption.png", bbox_inches="tight")
plt.close(fig)

# =========================================================
# CHART 7: EU adoption by age (Eurostat)
# =========================================================
fig, ax = plt.subplots(figsize=(8, 4.5))
ages = ["16-24", "25-54\n(est.)", "55-64\n(est.)", "65-74"]
pcts = [63.8, 36, 18, 7]
colors = ["#1F4E78", "#2E75B6", "#5B9BD5", "#9DC3E6"]
bars = ax.bar(ages, pcts, color=colors, edgecolor="white", linewidth=1.5)
ax.set_ylabel("% who used generative AI tools in 2025")
ax.set_title("EU Generative AI Use by Age — Eurostat 2025")
ax.set_ylim(0, 75)
for b, p in zip(bars, pcts):
    ax.text(b.get_x() + b.get_width()/2, b.get_height() + 1.5, f"{p}%", ha="center", fontweight="bold")
ax.axhline(32.7, color="#C00000", linestyle="--", linewidth=1, alpha=0.7)
ax.text(3.4, 34.2, "EU avg 32.7%", color="#C00000", fontsize=9, ha="right")
fig.tight_layout()
fig.savefig(OUTDIR / "chart7_eu_adoption_by_age.png", bbox_inches="tight")
plt.close(fig)

print("\nAll 7 charts saved.")
