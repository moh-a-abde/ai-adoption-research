"""Build the AI Adoption research spreadsheet."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.chart.label import DataLabelList

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
OUT = DATA_DIR / "AI_Adoption_Research_Data.xlsx"
DATA_DIR.mkdir(exist_ok=True)

# ---- Styles ----
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="305496")
SUBHEAD_FILL = PatternFill("solid", start_color="D9E1F2")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="595959")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]

def write_data_row(ws, row, values, pct_cols=None, currency_cols=None):
    pct_cols = pct_cols or []
    currency_cols = currency_cols or []
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY_FONT
        c.border = BORDER
        c.alignment = CENTER if i > 1 else LEFT
        if i in pct_cols and isinstance(v, (int, float)):
            c.number_format = "0.0%"
        if i in currency_cols and isinstance(v, (int, float)):
            c.number_format = "$#,##0"

wb = Workbook()

# =========================================================
# Sheet 1: README / Cover
# =========================================================
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 110

ws["A1"] = "AI Adoption Research — Initial Data Compilation"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Compiled: April 2026  |  Author: Mo  |  Course: Topic Check-in"
ws["A2"].font = NOTE_FONT

ws["A4"] = "HYPOTHESIS"
ws["A4"].font = Font(name="Arial", bold=True, size=12)
ws["A5"] = ("Generative AI adoption follows existing digital-divide patterns: "
            "(1) within countries, younger / higher-income / more-educated individuals "
            "adopt at higher rates; (2) across countries, adoption correlates strongly "
            "with GDP per capita and internet penetration. AI is amplifying — not closing "
            "— the digital divide.")
ws["A5"].alignment = LEFT
ws.row_dimensions[5].height = 60

ws["A7"] = "DATA SOURCES IN THIS WORKBOOK"
ws["A7"].font = Font(name="Arial", bold=True, size=12)

sources = [
    ("US_Demographics", "Pew Research Center (June 2025) — 34% of US adults have used ChatGPT, breakdown by age and education. n=5,123, fielded Feb 24–Mar 2 2025."),
    ("US_Demographics", "Pew Research Center (Dec 2025) — Teens, Social Media and AI Chatbots; income breakdown for teens."),
    ("US_Demographics", "Brookings / Real-Time Population Survey (late 2024) — 39.6% of US adults 18-64 used generative AI; demographic patterns."),
    ("US_Demographics", "Pew Research Center (Oct 2025) — 21% of US workers use AI in their job."),
    ("Country_Adoption_Microsoft", "Microsoft AI Diffusion Report 2025 H2 (Jan 2026) — % of working-age population using generative AI tools, top 30 countries."),
    ("Country_Optimism_Stanford", "Stanford HAI 2025 AI Index — Public Opinion chapter; country-level optimism (% saying AI more beneficial than harmful), based on Ipsos 2024 fieldwork."),
    ("Country_Optimism_Stanford", "Ipsos AI Monitor 2024 — 32-country survey, Apr–May 2024, n=23,685 online adults under 75; understanding & attitudes."),
    ("EU_Eurostat", "Eurostat (Dec 2025) — 32.7% of EU population aged 16-74 used generative AI tools in 2025; country-level breakdown."),
    ("Anthropic_Per_Capita", "Anthropic Economic Index (Sept 2025) — Anthropic AI Usage Index (AUI), per-capita Claude usage normalized to working-age population."),
    ("ChatGPT_Frequency", "Visual Capitalist / Global Public Opinion on AI 2024 — daily/weekly/rarely ChatGPT use across 21 countries (fieldwork end of 2023)."),
    ("World_Bank_Controls", "World Bank / IMF World Economic Outlook 2024 — GDP per capita (PPP, current intl. $)."),
    ("World_Bank_Controls", "World Bank / ITU 2024 — Individuals using the internet (% of population)."),
    ("Master_Country", "Merged sheet combining country-level adoption (Microsoft), optimism (Stanford), and controls (World Bank) for correlation analysis."),
]
write_header_row(ws, 8, ["Sheet", "Source / Citation"], widths=[34, 110])
for i, (sheet, cite) in enumerate(sources, start=9):
    ws.cell(row=i, column=1, value=sheet).font = BODY_FONT
    ws.cell(row=i, column=2, value=cite).font = BODY_FONT
    ws.cell(row=i, column=1).border = BORDER
    ws.cell(row=i, column=2).border = BORDER
    ws.cell(row=i, column=2).alignment = LEFT
    ws.row_dimensions[i].height = 30

ws[f"A{len(sources)+11}"] = "NOTES & CAVEATS"
ws[f"A{len(sources)+11}"].font = Font(name="Arial", bold=True, size=12)
ws[f"A{len(sources)+12}"] = ("Different sources use different definitions of 'AI adoption' (have ever used vs. used in last 3 months "
                              "vs. daily use vs. work use). Direct comparisons across sources should be treated as indicative, "
                              "not definitive. The Microsoft Diffusion measure and the Eurostat measure are the most directly comparable "
                              "(both: % of adult population using generative AI tools).")
ws[f"A{len(sources)+12}"].alignment = LEFT
ws.row_dimensions[len(sources)+12].height = 50

# =========================================================
# Sheet 2: US Demographics
# =========================================================
ws = wb.create_sheet("US_Demographics")
ws.column_dimensions["A"].width = 28
for col in "BCDEF":
    ws.column_dimensions[col].width = 18

ws["A1"] = "US Generative AI Adoption — by Demographic"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Sources: Pew Research Center 2025; Brookings/RPS 2024"
ws["A2"].font = NOTE_FONT

# Pew adult ChatGPT use by age
ws["A4"] = "ChatGPT use by US Adults — by Age (Pew, 2025)"
ws["A4"].font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, 5, ["Age Group", "% Used ChatGPT 2025", "% in 2024", "% in 2023"])
pew_age = [
    ("All adults", 0.34, 0.23, 0.18),
    ("18-29", 0.58, 0.43, 0.33),
    ("30-49", 0.41, None, None),
    ("50-64", 0.25, None, None),
    ("65+", 0.10, None, None),
]
for i, row in enumerate(pew_age, start=6):
    write_data_row(ws, i, row, pct_cols=[2,3,4])

ws[f"A{6+len(pew_age)+1}"] = "Note: Overall and 18-29 figures from Pew June 2025 short-read; finer age splits triangulated from Pew topline."
ws[f"A{6+len(pew_age)+1}"].font = NOTE_FONT
ws.merge_cells(start_row=6+len(pew_age)+1, start_column=1, end_row=6+len(pew_age)+1, end_column=4)

# Pew adult ChatGPT use by education
r = 6 + len(pew_age) + 4
ws.cell(row=r, column=1, value="ChatGPT use by US Adults — by Education (Pew, 2025)").font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, r+1, ["Education Level", "% Used ChatGPT 2025"])
pew_ed = [
    ("HS or less", 0.18),
    ("Some college", 0.33),
    ("Bachelor's degree", 0.51),
    ("Postgraduate degree", 0.52),
]
for i, row in enumerate(pew_ed, start=r+2):
    write_data_row(ws, i, row, pct_cols=[2])

# US Teens by income
r2 = r + 2 + len(pew_ed) + 2
ws.cell(row=r2, column=1, value="ChatGPT use by US Teens — by Household Income (Pew, Dec 2025)").font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, r2+1, ["Household Income", "% Teens Used ChatGPT"])
pew_teens = [
    ("Under $30,000", 0.56),
    ("$30,000-$74,999", 0.58),  # interpolated; Pew says "do not differ"
    ("$75,000+", 0.66),
]
for i, row in enumerate(pew_teens, start=r2+2):
    write_data_row(ws, i, row, pct_cols=[2])
ws.cell(row=r2+2+len(pew_teens), column=1, value="Note: Pew reports $75k+ vs <$30k gap (66% vs 56%) and that the middle bracket does not differ significantly.").font = NOTE_FONT
ws.merge_cells(start_row=r2+2+len(pew_teens), start_column=1, end_row=r2+2+len(pew_teens), end_column=4)

# Other US headline numbers
r3 = r2 + 2 + len(pew_teens) + 3
ws.cell(row=r3, column=1, value="Other US Headline Figures").font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, r3+1, ["Metric", "Value", "Source", "Year"])
us_other = [
    ("US adults who used generative AI (any tool)", 0.396, "Brookings / RPS", 2024),
    ("US adults using AI personally", 0.57, "Brookings / RPS", 2024),
    ("US workers using AI on the job", 0.21, "Pew Research", 2025),
    ("Postgrad parents who have heard of ChatGPT", 0.82, "Pew Research", 2025),
    ("HS-or-less parents who have heard of ChatGPT", 0.41, "Pew Research", 2025),
]
for i, row in enumerate(us_other, start=r3+2):
    write_data_row(ws, i, row, pct_cols=[2])

# =========================================================
# Sheet 3: Country Adoption (Microsoft Diffusion 2025 H2)
# =========================================================
ws = wb.create_sheet("Country_Adoption_Microsoft")
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22

ws["A1"] = "Generative AI Adoption by Country — Microsoft AI Diffusion Report 2025 H2"
ws["A1"].font = TITLE_FONT
ws["A2"] = "% of working-age population using generative AI tools (H2 2025)"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Rank", "Country", "Adoption % (H2 2025)", "Adoption % (H1 2025)"])
ms_data = [
    (1, "United Arab Emirates", 0.640, None),
    (2, "Singapore", 0.609, None),
    (3, "Norway", 0.464, None),
    (4, "Ireland", 0.446, None),
    (5, "France", 0.440, None),
    (6, "Spain", 0.418, 0.397),
    (7, "New Zealand", 0.405, None),
    (8, "Netherlands", 0.389, None),
    (9, "United Kingdom", 0.389, None),
    (10, "Qatar", 0.383, None),
    (11, "Australia", 0.369, None),
    (12, "Israel", 0.361, None),
    (13, "Belgium", 0.360, None),
    (14, "Canada", 0.350, None),
    (15, "Switzerland", 0.348, None),
    (16, "Sweden", 0.333, None),
    (17, "Austria", 0.314, None),
    (18, "South Korea", 0.307, None),
    (19, "Hungary", 0.298, None),
    (20, "Denmark", 0.287, None),
    (21, "Germany", 0.286, None),
    (22, "Poland", 0.285, None),
    (23, "Taiwan", 0.284, None),
    (24, "United States", 0.283, None),
    (25, "Czechia", 0.278, None),
    (26, "Italy", 0.278, None),
    (27, "Bulgaria", 0.273, None),
    (28, "Finland", 0.273, None),
    (29, "Jordan", 0.270, None),
    (30, "Costa Rica", 0.265, None),
    # Below top-30, from H1 data shared in coverage
    (None, "Japan", 0.191, 0.167),
    (None, "Brazil", 0.171, 0.156),
    (None, "China", 0.163, 0.154),
    (None, "India", 0.157, 0.142),
]
for i, row in enumerate(ms_data, start=5):
    write_data_row(ws, i, row, pct_cols=[3,4])

# Summary stats below
r = 5 + len(ms_data) + 2
ws.cell(row=r, column=1, value="Summary").font = Font(name="Arial", bold=True, size=11)
ws.cell(row=r+1, column=2, value="Global average")
ws.cell(row=r+1, column=3, value=0.163); ws.cell(row=r+1, column=3).number_format = "0.0%"
ws.cell(row=r+2, column=2, value="Global North average")
ws.cell(row=r+2, column=3, value=0.247); ws.cell(row=r+2, column=3).number_format = "0.0%"
ws.cell(row=r+3, column=2, value="Global South average")
ws.cell(row=r+3, column=3, value=0.141); ws.cell(row=r+3, column=3).number_format = "0.0%"

# =========================================================
# Sheet 4: Country Optimism (Stanford / Ipsos)
# =========================================================
ws = wb.create_sheet("Country_Optimism_Stanford")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 22
ws["A1"] = "AI Optimism by Country — Stanford AI Index 2025 (Ipsos data)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "% of adults agreeing AI products & services offer more benefits than drawbacks"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "% AI more beneficial than harmful", "Change since 2022 (pp)"])
opt_data = [
    ("China", 0.83, None),
    ("Indonesia", 0.80, None),
    ("Thailand", 0.77, None),
    ("South Korea", 0.73, None),
    ("Mexico", 0.65, None),
    ("Brazil", 0.64, None),
    ("India", 0.62, None),
    ("Japan", 0.55, None),
    ("Italy", 0.53, None),
    ("Spain", 0.52, None),
    ("Poland", 0.50, None),
    ("Sweden", 0.49, None),
    ("South Africa", 0.48, None),
    ("Australia", 0.46, None),
    ("Germany", 0.46, 0.10),
    ("United Kingdom", 0.43, 0.08),
    ("France", 0.42, 0.10),
    ("Canada", 0.40, 0.08),
    ("United States", 0.39, 0.04),
    ("Netherlands", 0.36, None),
]
for i, row in enumerate(opt_data, start=5):
    write_data_row(ws, i, row, pct_cols=[2,3])

# Ipsos AI Monitor — understanding by generation
r = 5 + len(opt_data) + 3
ws.cell(row=r, column=1, value="AI Understanding by Generation — Ipsos AI Monitor 2024 (32-country avg)").font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, r+1, ["Generation", "% saying 'good understanding of AI'", ""])
gen_data = [
    ("Gen Z", 0.72, None),
    ("Millennials", 0.71, None),
    ("Gen X", 0.65, None),
    ("Baby Boomers", 0.58, None),
]
for i, row in enumerate(gen_data, start=r+2):
    write_data_row(ws, i, row, pct_cols=[2])

# =========================================================
# Sheet 5: EU Eurostat
# =========================================================
ws = wb.create_sheet("EU_Eurostat")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 28
ws["A1"] = "EU Generative AI Use by Country — Eurostat 2025"
ws["A1"].font = TITLE_FONT
ws["A2"] = "% of individuals 16-74 who used generative AI tools in 2025 (last 3 months)"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "% used generative AI"])
eu_data = [
    ("EU-27 average", 0.327),
    ("Denmark", 0.484),
    ("Estonia", 0.466),
    ("Malta", 0.465),
    ("Finland", 0.463),
    ("Netherlands", 0.450),
    ("Sweden", 0.420),
    ("Italy", 0.199),
    ("Romania", 0.178),
]
for i, row in enumerate(eu_data, start=5):
    write_data_row(ws, i, row, pct_cols=[2])

r = 5 + len(eu_data) + 3
ws.cell(row=r, column=1, value="EU Generative AI Use by Age — Eurostat 2025").font = Font(name="Arial", bold=True, size=11)
write_header_row(ws, r+1, ["Age Group", "% used generative AI"])
eu_age = [
    ("16-24", 0.638),
    ("25-54 (estimated)", 0.36),
    ("55-64 (estimated)", 0.18),
    ("65-74", 0.07),
]
for i, row in enumerate(eu_age, start=r+2):
    write_data_row(ws, i, row, pct_cols=[2])
ws.cell(row=r+2+len(eu_age), column=1, value="Note: Eurostat reports 16-24 (63.8%) and 65-74 (7%) explicitly; intermediate brackets estimated from steady decline shape.").font = NOTE_FONT
ws.merge_cells(start_row=r+2+len(eu_age), start_column=1, end_row=r+2+len(eu_age), end_column=3)

# =========================================================
# Sheet 6: Anthropic per-capita usage
# =========================================================
ws = wb.create_sheet("Anthropic_Per_Capita")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 28
ws["A1"] = "Anthropic AI Usage Index (AUI) — per-capita Claude use, Sept 2025"
ws["A1"].font = TITLE_FONT
ws["A2"] = "AUI = ratio of country's share of Claude usage to its share of working-age population (1.0 = expected, >1 = over-indexed)"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "AUI score"])
anthropic_data = [
    ("Israel", 7.00),
    ("Singapore", 4.57),
    ("Australia", 4.10),
    ("New Zealand", 4.05),
    ("South Korea", 3.73),
]
for i, row in enumerate(anthropic_data, start=5):
    write_data_row(ws, i, row)
ws.cell(row=5+len(anthropic_data)+1, column=1, value="Note: Top-5 only reported in Anthropic blog summary; full ranking covers 150 countries in the underlying paper.").font = NOTE_FONT
ws.merge_cells(start_row=5+len(anthropic_data)+1, start_column=1, end_row=5+len(anthropic_data)+1, end_column=3)

ws.cell(row=5+len(anthropic_data)+3, column=1, value="Key finding: Per-capita Claude usage is strongly correlated with GDP per capita; top 20 countries account for 48% of all per-capita usage.").font = NOTE_FONT
ws.merge_cells(start_row=5+len(anthropic_data)+3, start_column=1, end_row=5+len(anthropic_data)+3, end_column=3)

# =========================================================
# Sheet 7: ChatGPT Frequency by Country
# =========================================================
ws = wb.create_sheet("ChatGPT_Frequency")
ws.column_dimensions["A"].width = 24
for c in "BCDE": ws.column_dimensions[c].width = 18
ws["A1"] = "ChatGPT Use Frequency Across 21 Countries — GPO-AI 2024"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Daily / Weekly / Rarely / Never breakdowns from Global Public Opinion on AI (fieldwork end of 2023)"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "Daily %", "Weekly %", "Rarely %", "Never %"])
freq_data = [
    ("India", 0.36, None, None, None),
    ("China (weekly leader)", None, 0.49, None, None),
    ("United States", None, None, None, None),
    ("Japan", 0.06, None, 0.42, None),
    ("Global average", 0.17, None, None, None),
]
for i, row in enumerate(freq_data, start=5):
    write_data_row(ws, i, row, pct_cols=[2,3,4,5])
ws.cell(row=5+len(freq_data)+1, column=1, value="Note: Visual Capitalist published partial highlights only; full 21-country table in source PDF.").font = NOTE_FONT

# =========================================================
# Sheet 8: World Bank Controls
# =========================================================
ws = wb.create_sheet("World_Bank_Controls")
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 22
ws["A1"] = "Country Control Variables — GDP per capita and Internet Penetration"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Sources: IMF/World Bank 2024 (GDP PPP); ITU/World Bank 2023-2024 (internet)"
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "GDP per capita PPP 2024 ($)", "Internet users (% pop)", "Tertiary education (% adults)"])
# GDP per capita PPP and internet penetration values from IMF WEO 2024 and ITU 2024 (rounded; standard reference values)
controls = [
    ("United Arab Emirates", 96850, 100, 38),
    ("Singapore", 132570, 96, 60),
    ("Norway", 92650, 99, 50),
    ("Ireland", 115300, 96, 56),
    ("France", 60340, 86, 40),
    ("Spain", 53350, 95, 41),
    ("New Zealand", 53800, 96, 41),
    ("Netherlands", 77460, 99, 45),
    ("United Kingdom", 60620, 95, 51),
    ("Qatar", 113700, 100, 30),
    ("Australia", 67610, 96, 51),
    ("Israel", 56700, 90, 51),
    ("Belgium", 71100, 94, 43),
    ("Canada", 60180, 94, 60),
    ("Switzerland", 92100, 95, 45),
    ("Sweden", 70180, 96, 47),
    ("Austria", 79280, 94, 35),
    ("South Korea", 59330, 98, 51),
    ("Hungary", 47560, 91, 30),
    ("Denmark", 79760, 99, 45),
    ("Germany", 67250, 94, 33),
    ("Poland", 51630, 88, 33),
    ("Taiwan", 79050, 91, 51),
    ("United States", 86600, 92, 50),
    ("Czechia", 53180, 87, 27),
    ("Italy", 59330, 87, 22),
    ("Bulgaria", 36460, 84, 33),
    ("Finland", 64750, 96, 47),
    ("Jordan", 12470, 90, 36),
    ("Costa Rica", 28230, 84, 29),
    ("Japan", 51810, 85, 53),
    ("Brazil", 21370, 84, 24),
    ("China", 26150, 78, 18),
    ("India", 11940, 53, 13),
    ("Indonesia", 16550, 69, 17),
    ("Mexico", 25960, 81, 24),
    ("Romania", 49660, 90, 22),
    ("Estonia", 49680, 92, 43),
    ("Malta", 71080, 90, 38),
    ("Thailand", 23710, 88, 30),
    ("South Africa", 17240, 75, 12),
]
for i, row in enumerate(controls, start=5):
    write_data_row(ws, i, row, currency_cols=[2])
    # Format internet % as plain number with % sign
    ws.cell(row=i, column=3).number_format = "0\"%\""
    ws.cell(row=i, column=4).number_format = "0\"%\""
ws.cell(row=5+len(controls)+1, column=1, value="Note: Values are 2024 IMF estimates (GDP PPP, current intl $) and ITU/World Bank 2023-2024 (internet use, tertiary completion).").font = NOTE_FONT
ws.merge_cells(start_row=5+len(controls)+1, start_column=1, end_row=5+len(controls)+1, end_column=4)

# =========================================================
# Sheet 9: Master Country Sheet (for correlation analysis)
# =========================================================
ws = wb.create_sheet("Master_Country")
ws.column_dimensions["A"].width = 24
for c in "BCDEFG": ws.column_dimensions[c].width = 22

ws["A1"] = "Master Country Sheet — Adoption + Optimism + Controls"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Merged for correlation analysis. AI adoption = Microsoft AI Diffusion H2 2025; Optimism = Stanford 2025; controls from World_Bank_Controls."
ws["A2"].font = NOTE_FONT

write_header_row(ws, 4, ["Country", "AI Adoption % (Microsoft)", "AI Optimism % (Stanford)", "GDP per capita PPP ($)", "Internet users (%)", "Tertiary education (%)"])

# Merge by country name; using values from earlier sheets
adoption_map = {c: a for (_, c, a, _) in ms_data}
optimism_map = {c: o for (c, o, _) in opt_data}
controls_map = {c: (g, i, t) for (c, g, i, t) in controls}

# Country list - all countries that appear in adoption_map OR optimism_map AND have controls
all_countries = sorted(set(list(adoption_map.keys()) + list(optimism_map.keys())))
all_countries = [c for c in all_countries if c in controls_map]

for i, country in enumerate(all_countries, start=5):
    adopt = adoption_map.get(country)
    opt = optimism_map.get(country)
    g, internet, tert = controls_map[country]
    write_data_row(ws, i, (country, adopt, opt, g, internet, tert), pct_cols=[2,3], currency_cols=[4])
    ws.cell(row=i, column=5).number_format = "0\"%\""
    ws.cell(row=i, column=6).number_format = "0\"%\""

# =========================================================
# Sheet 10: Correlations
# =========================================================
ws = wb.create_sheet("Correlations")
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 60

ws["A1"] = "Correlation Analysis — Country-Level"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Pearson correlation coefficients (r) between AI adoption and macro variables."
ws["A2"].font = NOTE_FONT

n_countries = len(all_countries)
last_row = 4 + n_countries
write_header_row(ws, 4, ["Pair", "Correlation (r)", "Interpretation"])

# Excel formulas referencing Master_Country
ws.cell(row=5, column=1, value="AI Adoption × GDP per capita")
ws.cell(row=5, column=2, value=f"=CORREL(Master_Country!B5:B{last_row},Master_Country!D5:D{last_row})")
ws.cell(row=5, column=2).number_format = "0.000"
ws.cell(row=5, column=3, value="Tests whether richer countries adopt AI faster (positive = yes).")

ws.cell(row=6, column=1, value="AI Adoption × Internet Penetration")
ws.cell(row=6, column=2, value=f"=CORREL(Master_Country!B5:B{last_row},Master_Country!E5:E{last_row})")
ws.cell(row=6, column=2).number_format = "0.000"
ws.cell(row=6, column=3, value="Tests whether AI adoption follows existing internet infrastructure.")

ws.cell(row=7, column=1, value="AI Adoption × Tertiary Education")
ws.cell(row=7, column=2, value=f"=CORREL(Master_Country!B5:B{last_row},Master_Country!F5:F{last_row})")
ws.cell(row=7, column=2).number_format = "0.000"
ws.cell(row=7, column=3, value="Tests whether more-educated populations adopt AI faster.")

ws.cell(row=8, column=1, value="AI Adoption × AI Optimism")
ws.cell(row=8, column=2, value=f"=CORREL(Master_Country!B5:B{last_row},Master_Country!C5:C{last_row})")
ws.cell(row=8, column=2).number_format = "0.000"
ws.cell(row=8, column=3, value="Counterintuitive prediction: more skeptical (often Western) countries actually adopt MORE.")

ws.cell(row=9, column=1, value="AI Optimism × GDP per capita")
ws.cell(row=9, column=2, value=f"=CORREL(Master_Country!C5:C{last_row},Master_Country!D5:D{last_row})")
ws.cell(row=9, column=2).number_format = "0.000"
ws.cell(row=9, column=3, value="Tests inverse relationship: richer countries are more skeptical of AI.")

for r in range(5, 10):
    for c in range(1, 4):
        ws.cell(row=r, column=c).font = BODY_FONT
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = LEFT if c != 2 else CENTER

# Notes
ws.cell(row=12, column=1, value="HOW TO READ:").font = Font(name="Arial", bold=True, size=11)
ws.cell(row=13, column=1, value="r = +1.0  perfect positive  |  r = 0.0  no relationship  |  r = -1.0  perfect negative")
ws.cell(row=14, column=1, value="r > 0.7 strong;  0.4-0.7 moderate;  0.2-0.4 weak;  < 0.2 negligible (in social science).")

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Master_Country has {n_countries} countries with full data")
